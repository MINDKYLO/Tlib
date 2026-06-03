from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, Response, send_from_directory)
import psycopg2, psycopg2.extras
import qrcode, io, base64, csv, os, smtplib, hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date
from functools import wraps
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'it-equipment-secret-key-2024'
ALLOWED = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

_ON_VERCEL    = os.environ.get('VERCEL') or not os.access('.', os.W_OK)
UPLOAD_FOLDER = '/tmp/uploads' if _ON_VERCEL else os.path.join('static', 'uploads')

DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://localhost/itborrow')
# Railway/Heroku use "postgres://" prefix; psycopg2 requires "postgresql://"
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)


# ── DB HELPERS ─────────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DATABASE_URL,
                            cursor_factory=psycopg2.extras.RealDictCursor)


def _to_str(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(v, date):
        return v.isoformat()
    return v


def row_to_dict(row):
    if row is None:
        return None
    return {k: _to_str(v) for k, v in dict(row).items()}


def rows_to_dicts(rows):
    return [row_to_dict(r) for r in rows]


def db_fetchone(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    return row_to_dict(cur.fetchone())


def db_fetchall(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    return rows_to_dicts(cur.fetchall())


def db_execute(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    return cur


def init_db():
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    conn = get_db()
    cur  = conn.cursor()

    cur.execute('''CREATE TABLE IF NOT EXISTS users (
        id         SERIAL PRIMARY KEY,
        username   TEXT UNIQUE NOT NULL,
        password   TEXT NOT NULL,
        name       TEXT NOT NULL,
        role       TEXT DEFAULT 'user',
        department TEXT,
        email      TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS equipment (
        id            SERIAL PRIMARY KEY,
        name          TEXT NOT NULL,
        category      TEXT NOT NULL,
        serial_number TEXT UNIQUE,
        brand         TEXT,
        model         TEXT,
        status        TEXT DEFAULT 'available',
        description   TEXT,
        image         TEXT,
        added_date    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS borrows (
        id           SERIAL PRIMARY KEY,
        equipment_id INTEGER NOT NULL REFERENCES equipment(id),
        user_id      INTEGER NOT NULL REFERENCES users(id),
        borrow_date  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        due_date     DATE,
        return_date  TIMESTAMP,
        status       TEXT DEFAULT 'borrowed',
        notes        TEXT
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS email_settings (
        id            SERIAL PRIMARY KEY,
        smtp_host     TEXT DEFAULT '',
        smtp_port     INTEGER DEFAULT 587,
        smtp_user     TEXT DEFAULT '',
        smtp_password TEXT DEFAULT '',
        smtp_from     TEXT DEFAULT '',
        enabled       SMALLINT DEFAULT 0
    )''')

    cur.execute('''CREATE TABLE IF NOT EXISTS categories (
        id         SERIAL PRIMARY KEY,
        name       TEXT UNIQUE NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    default_cats = ['Notebook','Monitor','Mobile','Tablet','Camera','Printer','Network','Storage','Peripheral','Other']
    for cat in default_cats:
        cur.execute('INSERT INTO categories (name) VALUES (%s) ON CONFLICT (name) DO NOTHING', (cat,))

    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS email TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS phone TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS id_number TEXT")
    cur.execute("ALTER TABLE equipment ADD COLUMN IF NOT EXISTS image TEXT")

    cur.execute('SELECT COUNT(*) as c FROM email_settings')
    if not cur.fetchone()['c']:
        cur.execute("INSERT INTO email_settings (smtp_host, smtp_port) VALUES ('', 587)")

    admin_pw = hashlib.sha256('admin123'.encode()).hexdigest()
    cur.execute('''INSERT INTO users (username, password, name, role, department)
                   VALUES (%s, %s, %s, %s, %s)
                   ON CONFLICT (username) DO NOTHING''',
                ('admin', admin_pw, 'ผู้ดูแลระบบ', 'admin', 'ฝ่าย IT'))

    sample = [
        ('MacBook Pro 14"',  'Notebook', 'NB-001', 'Apple', 'MacBook Pro M2', 'มี 16GB RAM / 512GB SSD'),
        ('Dell Monitor 27"', 'Monitor',  'MN-001', 'Dell',  'U2722D 4K',      'จอ 4K IPS 27 นิ้ว'),
        ('iPhone 14 Pro',    'Mobile',   'MB-001', 'Apple', 'iPhone 14 Pro',  '256GB Deep Purple'),
        ('Canon EOS R5',     'Camera',   'CM-001', 'Canon', 'EOS R5',         'กล้อง Mirrorless 45MP'),
        ('iPad Pro 12.9"',   'Tablet',   'TB-001', 'Apple', 'iPad Pro M2',    '256GB WiFi+Cellular'),
        ('HP LaserJet Pro',  'Printer',  'PR-001', 'HP',    'LaserJet M404n', 'เครื่องพิมพ์ขาวดำ'),
    ]
    for s in sample:
        cur.execute('''INSERT INTO equipment (name, category, serial_number, brand, model, description)
                       VALUES (%s, %s, %s, %s, %s, %s)
                       ON CONFLICT (serial_number) DO NOTHING''', s)

    conn.commit()
    conn.close()


# ── AUTH DECORATORS ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('ไม่มีสิทธิ์เข้าถึงหน้านี้', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ── EMAIL HELPERS ──────────────────────────────────────────────────────────────

def get_email_cfg():
    conn = get_db()
    row  = db_fetchone(conn, 'SELECT * FROM email_settings LIMIT 1')
    conn.close()
    return row or {}


def send_email(to_email, subject, html_body):
    if not to_email:
        return False
    cfg = get_email_cfg()
    if not cfg.get('enabled') or not cfg.get('smtp_host'):
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = cfg.get('smtp_from') or cfg.get('smtp_user', '')
        msg['To']      = to_email
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        with smtplib.SMTP(cfg['smtp_host'], int(cfg['smtp_port']), timeout=10) as s:
            s.ehlo()
            s.starttls()
            s.login(cfg['smtp_user'], cfg['smtp_password'])
            s.sendmail(msg['From'], [to_email], msg.as_string())
        return True
    except Exception as e:
        print(f'[Email Error] {e}')
        return False


def email_borrow(user_email, user_name, eq_name, due_date):
    html = f"""
    <div style="font-family:sans-serif;max-width:500px;margin:auto;padding:20px">
      <div style="background:#2563eb;color:#fff;padding:16px 20px;border-radius:8px 8px 0 0">
        <h2 style="margin:0">&#x2705; ยืมอุปกรณ์สำเร็จ</h2>
      </div>
      <div style="border:1px solid #e2e8f0;border-top:none;padding:20px;border-radius:0 0 8px 8px">
        <p>สวัสดีคุณ <strong>{user_name}</strong></p>
        <p>คุณได้ทำการยืม <strong>{eq_name}</strong> เรียบร้อยแล้ว</p>
        {'<p>&#x1F4C5; กำหนดคืน: <strong>' + str(due_date) + '</strong></p>' if due_date else ''}
        <p style="color:#64748b;font-size:0.85em">หากมีข้อสงสัยกรุณาติดต่อฝ่าย IT</p>
      </div>
    </div>"""
    return send_email(user_email, f'[IT Borrow] ยืมอุปกรณ์: {eq_name}', html)


def email_overdue(user_email, user_name, eq_name, due_date):
    html = f"""
    <div style="font-family:sans-serif;max-width:500px;margin:auto;padding:20px">
      <div style="background:#dc2626;color:#fff;padding:16px 20px;border-radius:8px 8px 0 0">
        <h2 style="margin:0">&#x26A0;&#xFE0F; อุปกรณ์เกินกำหนดคืน</h2>
      </div>
      <div style="border:1px solid #fecaca;border-top:none;padding:20px;border-radius:0 0 8px 8px;background:#fff5f5">
        <p>สวัสดีคุณ <strong>{user_name}</strong></p>
        <p>อุปกรณ์ <strong>{eq_name}</strong> ที่คุณยืมไว้เกินกำหนดคืนแล้ว</p>
        <p>&#x1F4C5; กำหนดคืนเดิม: <strong style="color:#dc2626">{due_date}</strong></p>
        <p>กรุณาคืนอุปกรณ์โดยเร็วที่สุด</p>
      </div>
    </div>"""
    return send_email(user_email, f'[IT Borrow] &#x26A0;&#xFE0F; เกินกำหนดคืน: {eq_name}', html)


# ── QR ─────────────────────────────────────────────────────────────────────────

def generate_qr(data):
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED


# ── AUTH ───────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = hashlib.sha256(request.form['password'].encode()).hexdigest()
        conn = get_db()
        user = db_fetchone(conn, 'SELECT * FROM users WHERE username=%s AND password=%s',
                           (username, password))
        conn.close()
        if user:
            session.update({'user_id': user['id'], 'username': user['username'],
                            'name': user['name'], 'role': user['role']})
            flash(f'ยินดีต้อนรับ คุณ{user["name"]}', 'success')
            return redirect(url_for('dashboard'))
        flash('ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── DASHBOARD ──────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    conn      = get_db()
    today     = date.today()
    today_str = today.isoformat()

    total     = db_fetchone(conn, "SELECT COUNT(*) as c FROM equipment")['c']
    available = db_fetchone(conn, "SELECT COUNT(*) as c FROM equipment WHERE status='available'")['c']
    borrowed  = db_fetchone(conn, "SELECT COUNT(*) as c FROM equipment WHERE status='borrowed'")['c']

    overdue = db_fetchall(conn, '''
        SELECT b.*, e.name as eq_name, u.name as user_name, u.email as user_email
        FROM borrows b
        JOIN equipment e ON b.equipment_id=e.id
        JOIN users u ON b.user_id=u.id
        WHERE b.status='borrowed' AND b.due_date IS NOT NULL AND b.due_date < CURRENT_DATE
        ORDER BY b.due_date ASC
    ''')

    recent = db_fetchall(conn, '''
        SELECT b.*, e.name as eq_name, e.category, u.name as user_name
        FROM borrows b
        JOIN equipment e ON b.equipment_id=e.id
        JOIN users u ON b.user_id=u.id
        ORDER BY b.borrow_date DESC LIMIT 8
    ''')

    my_borrows = db_fetchall(conn, '''
        SELECT b.*, e.name as eq_name, e.category, e.brand, e.image
        FROM borrows b JOIN equipment e ON b.equipment_id=e.id
        WHERE b.user_id=%s AND b.status='borrowed'
        ORDER BY b.borrow_date DESC
    ''', (session['user_id'],))

    conn.close()
    return render_template('dashboard.html',
                           total=total, available=available, borrowed=borrowed,
                           overdue=overdue, recent=recent, my_borrows=my_borrows,
                           today=today_str)


# ── EQUIPMENT ──────────────────────────────────────────────────────────────────

@app.route('/equipment')
@login_required
def equipment_list():
    search   = request.args.get('search', '')
    category = request.args.get('category', '')
    status   = request.args.get('status', '')
    conn     = get_db()
    query    = 'SELECT * FROM equipment WHERE 1=1'
    params   = []
    if search:
        query += ' AND (name ILIKE %s OR serial_number ILIKE %s OR brand ILIKE %s OR model ILIKE %s)'
        params.extend([f'%{search}%'] * 4)
    if category:
        query += ' AND category=%s'; params.append(category)
    if status:
        query += ' AND status=%s'; params.append(status)
    query += ' ORDER BY added_date DESC'
    equipment  = db_fetchall(conn, query, params)
    categories = db_fetchall(conn, 'SELECT DISTINCT category FROM equipment ORDER BY category')
    conn.close()
    return render_template('equipment_list.html',
                           equipment=equipment, categories=categories,
                           search=search, selected_category=category, selected_status=status)


@app.route('/equipment/add', methods=['GET', 'POST'])
@admin_required
def add_equipment():
    if request.method == 'POST':
        name   = request.form['name'];     category = request.form['category']
        serial = request.form['serial_number']; brand = request.form['brand']
        model  = request.form['model'];    desc     = request.form['description']
        conn   = get_db()
        try:
            cur = db_execute(conn, '''INSERT INTO equipment
                             (name, category, serial_number, brand, model, description)
                             VALUES (%s,%s,%s,%s,%s,%s) RETURNING id''',
                             (name, category, serial, brand, model, desc))
            eid = cur.fetchone()['id']
            conn.commit()
            f = request.files.get('image')
            if f and f.filename and allowed_file(f.filename):
                ext   = secure_filename(f.filename).rsplit('.', 1)[1].lower()
                fname = f'eq_{eid}.{ext}'
                f.save(os.path.join(UPLOAD_FOLDER, fname))
                db_execute(conn, 'UPDATE equipment SET image=%s WHERE id=%s', (fname, eid))
                conn.commit()
            flash('เพิ่มอุปกรณ์สำเร็จ', 'success')
            return redirect(url_for('equipment_list'))
        except psycopg2.IntegrityError:
            flash('Serial Number นี้มีอยู่แล้วในระบบ', 'danger')
        finally:
            conn.close()
    conn = get_db()
    categories = db_fetchall(conn, 'SELECT name FROM categories ORDER BY name')
    conn.close()
    return render_template('add_equipment.html', categories=categories)


@app.route('/equipment/<int:eid>')
@login_required
def equipment_detail(eid):
    conn = get_db()
    eq   = db_fetchone(conn, 'SELECT * FROM equipment WHERE id=%s', (eid,))
    if not eq:
        flash('ไม่พบอุปกรณ์', 'danger')
        conn.close()
        return redirect(url_for('equipment_list'))
    history = db_fetchall(conn, '''
        SELECT b.*, u.name as user_name, u.department
        FROM borrows b JOIN users u ON b.user_id=u.id
        WHERE b.equipment_id=%s ORDER BY b.borrow_date DESC
    ''', (eid,))
    current_borrow = db_fetchone(conn, '''
        SELECT b.*, u.name as user_name
        FROM borrows b JOIN users u ON b.user_id=u.id
        WHERE b.equipment_id=%s AND b.status='borrowed'
    ''', (eid,))
    qr_data  = f"ID:{eid}|ชื่อ:{eq['name']}|Serial:{eq['serial_number']}|หมวด:{eq['category']}"
    qr_image = generate_qr(qr_data)
    conn.close()
    return render_template('equipment_detail.html',
                           eq=eq, history=history, current_borrow=current_borrow,
                           qr_image=qr_image)


@app.route('/equipment/<int:eid>/edit', methods=['GET', 'POST'])
@admin_required
def edit_equipment(eid):
    conn = get_db()
    eq   = db_fetchone(conn, 'SELECT * FROM equipment WHERE id=%s', (eid,))
    if not eq:
        conn.close()
        return redirect(url_for('equipment_list'))
    if request.method == 'POST':
        name   = request.form['name'];     category = request.form['category']
        serial = request.form['serial_number']; brand = request.form['brand']
        model  = request.form['model'];    desc     = request.form['description']
        try:
            db_execute(conn, '''UPDATE equipment SET name=%s, category=%s, serial_number=%s,
                                brand=%s, model=%s, description=%s WHERE id=%s''',
                       (name, category, serial, brand, model, desc, eid))
            f = request.files.get('image')
            if f and f.filename and allowed_file(f.filename):
                ext   = secure_filename(f.filename).rsplit('.', 1)[1].lower()
                fname = f'eq_{eid}.{ext}'
                f.save(os.path.join(UPLOAD_FOLDER, fname))
                db_execute(conn, 'UPDATE equipment SET image=%s WHERE id=%s', (fname, eid))
            conn.commit()
            flash('แก้ไขข้อมูลอุปกรณ์สำเร็จ', 'success')
        except Exception:
            flash('เกิดข้อผิดพลาด', 'danger')
        finally:
            conn.close()
        return redirect(url_for('equipment_detail', eid=eid))
    categories = db_fetchall(conn, 'SELECT name FROM categories ORDER BY name')
    conn.close()
    return render_template('edit_equipment.html', eq=eq, categories=categories)


@app.route('/equipment/<int:eid>/delete', methods=['POST'])
@admin_required
def delete_equipment(eid):
    conn = get_db()
    eq   = db_fetchone(conn, 'SELECT * FROM equipment WHERE id=%s', (eid,))
    if eq and eq['status'] == 'available':
        if eq['image']:
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, eq['image']))
            except Exception:
                pass
        db_execute(conn, 'DELETE FROM equipment WHERE id=%s', (eid,))
        conn.commit()
        flash('ลบอุปกรณ์สำเร็จ', 'success')
    else:
        flash('ไม่สามารถลบอุปกรณ์ที่กำลังถูกยืมอยู่ได้', 'danger')
    conn.close()
    return redirect(url_for('equipment_list'))


# ── BORROW / RETURN ────────────────────────────────────────────────────────────

@app.route('/borrow/<int:eid>', methods=['POST'])
@login_required
def borrow(eid):
    conn = get_db()
    eq   = db_fetchone(conn, 'SELECT * FROM equipment WHERE id=%s', (eid,))
    if not eq or eq['status'] != 'available':
        flash('อุปกรณ์ไม่พร้อมให้ยืม', 'danger')
        conn.close()
        return redirect(url_for('equipment_list'))
    due_date = request.form.get('due_date') or None
    notes    = request.form.get('notes', '')
    db_execute(conn, 'INSERT INTO borrows (equipment_id, user_id, due_date, notes) VALUES (%s,%s,%s,%s)',
               (eid, session['user_id'], due_date, notes))
    db_execute(conn, "UPDATE equipment SET status='borrowed' WHERE id=%s", (eid,))
    conn.commit()
    user = db_fetchone(conn, 'SELECT * FROM users WHERE id=%s', (session['user_id'],))
    conn.close()
    if user and user.get('email'):
        email_borrow(user['email'], user['name'], eq['name'], due_date)
    flash(f'ยืม "{eq["name"]}" สำเร็จ', 'success')
    return redirect(url_for('dashboard'))


@app.route('/return/<int:bid>', methods=['POST'])
@login_required
def return_equipment(bid):
    conn       = get_db()
    borrow_row = db_fetchone(conn, '''
        SELECT b.*, e.name as eq_name
        FROM borrows b JOIN equipment e ON b.equipment_id=e.id WHERE b.id=%s
    ''', (bid,))
    if not borrow_row:
        flash('ไม่พบข้อมูลการยืม', 'danger')
        conn.close()
        return redirect(url_for('dashboard'))
    if borrow_row['user_id'] != session['user_id'] and session.get('role') != 'admin':
        flash('ไม่มีสิทธิ์คืนอุปกรณ์ชิ้นนี้', 'danger')
        conn.close()
        return redirect(url_for('dashboard'))
    db_execute(conn, "UPDATE borrows SET return_date=CURRENT_TIMESTAMP, status='returned' WHERE id=%s", (bid,))
    db_execute(conn, "UPDATE equipment SET status='available' WHERE id=%s", (borrow_row['equipment_id'],))
    conn.commit()
    conn.close()
    flash(f'คืน "{borrow_row["eq_name"]}" สำเร็จ', 'success')
    return redirect(url_for('dashboard'))


# ── HISTORY + EXPORT ───────────────────────────────────────────────────────────

@app.route('/history')
@login_required
def history():
    today = date.today().isoformat()
    conn  = get_db()
    if session.get('role') == 'admin':
        borrows = db_fetchall(conn, '''
            SELECT b.*, e.name as eq_name, e.category, u.name as user_name, u.department
            FROM borrows b
            JOIN equipment e ON b.equipment_id=e.id
            JOIN users u ON b.user_id=u.id
            ORDER BY b.borrow_date DESC
        ''')
    else:
        borrows = db_fetchall(conn, '''
            SELECT b.*, e.name as eq_name, e.category, u.name as user_name, u.department
            FROM borrows b
            JOIN equipment e ON b.equipment_id=e.id
            JOIN users u ON b.user_id=u.id
            WHERE b.user_id=%s ORDER BY b.borrow_date DESC
        ''', (session['user_id'],))
    conn.close()
    return render_template('history.html', borrows=borrows, today=today)


@app.route('/history/export')
@login_required
def export_history():
    fmt   = request.args.get('fmt', 'csv')
    today = date.today().isoformat()
    conn  = get_db()
    if session.get('role') == 'admin':
        rows = db_fetchall(conn, '''
            SELECT b.id, e.name as eq_name, e.category, e.serial_number,
                   u.name as user_name, u.department,
                   b.borrow_date, b.due_date, b.return_date, b.status, b.notes
            FROM borrows b
            JOIN equipment e ON b.equipment_id=e.id
            JOIN users u ON b.user_id=u.id
            ORDER BY b.borrow_date DESC
        ''')
    else:
        rows = db_fetchall(conn, '''
            SELECT b.id, e.name as eq_name, e.category, e.serial_number,
                   u.name as user_name, u.department,
                   b.borrow_date, b.due_date, b.return_date, b.status, b.notes
            FROM borrows b
            JOIN equipment e ON b.equipment_id=e.id
            JOIN users u ON b.user_id=u.id
            WHERE b.user_id=%s ORDER BY b.borrow_date DESC
        ''', (session['user_id'],))
    conn.close()

    headers_th = ['#', 'อุปกรณ์', 'หมวด', 'Serial', 'ผู้ยืม', 'แผนก',
                  'วันที่ยืม', 'ครบกำหนด', 'วันที่คืน', 'สถานะ', 'หมายเหตุ']

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ประวัติการยืม'
        hdr_fill = PatternFill('solid', fgColor='2563EB')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        for col, h in enumerate(headers_th, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        status_map = {'borrowed': 'ยืมอยู่', 'returned': 'คืนแล้ว'}
        for r, row in enumerate(rows, 2):
            vals = [row['id'], row['eq_name'], row['category'], row['serial_number'],
                    row['user_name'], row['department'],
                    row['borrow_date'][:16] if row['borrow_date'] else '',
                    row['due_date'] or '',
                    row['return_date'][:16] if row['return_date'] else '',
                    status_map.get(row['status'], row['status']),
                    row['notes'] or '']
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                if row['status'] == 'borrowed' and row['due_date'] and row['due_date'] < today:
                    cell.fill = PatternFill('solid', fgColor='FEE2E2')

        widths = [5, 25, 12, 14, 18, 14, 18, 12, 18, 10, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(buf.getvalue(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': f'attachment; filename=borrow_history_{today}.xlsx'})

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers_th)
    status_map = {'borrowed': 'ยืมอยู่', 'returned': 'คืนแล้ว'}
    for row in rows:
        writer.writerow([
            row['id'], row['eq_name'], row['category'], row['serial_number'],
            row['user_name'], row['department'],
            row['borrow_date'][:16] if row['borrow_date'] else '',
            row['due_date'] or '',
            row['return_date'][:16] if row['return_date'] else '',
            status_map.get(row['status'], row['status']),
            row['notes'] or ''
        ])
    output.seek(0)
    return Response('﻿' + output.getvalue(),
                    mimetype='text/csv; charset=utf-8-sig',
                    headers={'Content-Disposition': f'attachment; filename=borrow_history_{today}.csv'})


# ── USERS ──────────────────────────────────────────────────────────────────────

@app.route('/users')
@admin_required
def user_list():
    search = request.args.get('search', '').strip()
    conn   = get_db()
    if search:
        users = db_fetchall(conn, '''
            SELECT * FROM users
            WHERE name ILIKE %s OR username ILIKE %s
               OR department ILIKE %s OR id_number ILIKE %s
            ORDER BY created_at DESC
        ''', [f'%{search}%'] * 4)
    else:
        users = db_fetchall(conn, 'SELECT * FROM users ORDER BY created_at DESC')
    conn.close()
    return render_template('users.html', users=users, search=search)


@app.route('/users/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username   = request.form['username']
        password   = hashlib.sha256(request.form['password'].encode()).hexdigest()
        name       = request.form['name'];       role       = request.form['role']
        department = request.form['department']; email      = request.form.get('email', '')
        phone      = request.form.get('phone', ''); id_number = request.form.get('id_number', '')
        conn = get_db()
        try:
            db_execute(conn, '''INSERT INTO users
                         (username, password, name, role, department, email, phone, id_number)
                         VALUES (%s,%s,%s,%s,%s,%s,%s,%s)''',
                       (username, password, name, role, department, email, phone, id_number))
            conn.commit()
            flash('เพิ่มผู้ใช้สำเร็จ', 'success')
            return redirect(url_for('user_list'))
        except psycopg2.IntegrityError:
            flash('ชื่อผู้ใช้นี้มีอยู่แล้ว', 'danger')
        finally:
            conn.close()
    return render_template('add_user.html')


@app.route('/users/<int:uid>/edit', methods=['GET', 'POST'])
@admin_required
def edit_user(uid):
    conn = get_db()
    user = db_fetchone(conn, 'SELECT * FROM users WHERE id=%s', (uid,))
    if not user:
        conn.close()
        flash('ไม่พบผู้ใช้', 'danger')
        return redirect(url_for('user_list'))
    if request.method == 'POST':
        name       = request.form['name']
        department = request.form['department']
        email      = request.form.get('email', '')
        phone      = request.form.get('phone', '')
        id_number  = request.form.get('id_number', '')
        role       = request.form['role']
        db_execute(conn, '''UPDATE users SET name=%s, department=%s, email=%s, phone=%s,
                            id_number=%s, role=%s WHERE id=%s''',
                   (name, department, email, phone, id_number, role, uid))
        conn.commit()
        conn.close()
        flash(f'อัพเดทข้อมูล {name} สำเร็จ', 'success')
        return redirect(url_for('user_list'))
    conn.close()
    return render_template('edit_user.html', user=user)


@app.route('/users/<int:uid>/delete', methods=['POST'])
@admin_required
def delete_user(uid):
    if uid == session['user_id']:
        flash('ไม่สามารถลบบัญชีของตัวเองได้', 'danger')
        return redirect(url_for('user_list'))
    conn = get_db()
    db_execute(conn, 'DELETE FROM users WHERE id=%s', (uid,))
    conn.commit()
    conn.close()
    flash('ลบผู้ใช้สำเร็จ', 'success')
    return redirect(url_for('user_list'))


# ── PROFILE ────────────────────────────────────────────────────────────────────

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    conn = get_db()
    user = db_fetchone(conn, 'SELECT * FROM users WHERE id=%s', (session['user_id'],))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'update_info':
            name       = request.form['name']
            department = request.form['department']
            email      = request.form.get('email', '')
            phone      = request.form.get('phone', '')
            id_number  = request.form.get('id_number', '')
            db_execute(conn, '''UPDATE users SET name=%s, department=%s, email=%s,
                                phone=%s, id_number=%s WHERE id=%s''',
                       (name, department, email, phone, id_number, session['user_id']))
            conn.commit()
            session['name'] = name
            flash('อัพเดทข้อมูลสำเร็จ', 'success')

        elif action == 'change_password':
            old_pw  = hashlib.sha256(request.form['old_password'].encode()).hexdigest()
            new_pw  = hashlib.sha256(request.form['new_password'].encode()).hexdigest()
            confirm = hashlib.sha256(request.form['confirm_password'].encode()).hexdigest()
            if old_pw != user['password']:
                flash('รหัสผ่านเดิมไม่ถูกต้อง', 'danger')
            elif new_pw != confirm:
                flash('รหัสผ่านใหม่ไม่ตรงกัน', 'danger')
            else:
                db_execute(conn, 'UPDATE users SET password=%s WHERE id=%s',
                           (new_pw, session['user_id']))
                conn.commit()
                flash('เปลี่ยนรหัสผ่านสำเร็จ', 'success')

        conn.close()
        return redirect(url_for('profile'))

    my_borrows = db_fetchall(conn, '''
        SELECT b.*, e.name as eq_name, e.category
        FROM borrows b JOIN equipment e ON b.equipment_id=e.id
        WHERE b.user_id=%s ORDER BY b.borrow_date DESC LIMIT 10
    ''', (session['user_id'],))
    conn.close()
    return render_template('profile.html', user=user, my_borrows=my_borrows)


# ── EMAIL SETTINGS ─────────────────────────────────────────────────────────────

@app.route('/admin/email-settings', methods=['GET', 'POST'])
@admin_required
def email_settings_page():
    conn = get_db()
    cfg  = db_fetchone(conn, 'SELECT * FROM email_settings LIMIT 1')

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save':
            host     = request.form['smtp_host']
            port     = int(request.form.get('smtp_port', 587))
            user     = request.form['smtp_user']
            password = request.form['smtp_password']
            from_    = request.form['smtp_from']
            enabled  = 1 if request.form.get('enabled') else 0
            db_execute(conn, '''UPDATE email_settings SET smtp_host=%s, smtp_port=%s, smtp_user=%s,
                                smtp_password=%s, smtp_from=%s, enabled=%s WHERE id=%s''',
                       (host, port, user, password, from_, enabled, cfg['id']))
            conn.commit()
            flash('บันทึกการตั้งค่า Email สำเร็จ', 'success')

        elif action == 'test':
            test_to = request.form.get('test_email', '')
            ok = send_email(test_to, '[IT Borrow] ทดสอบการส่ง Email',
                            '<p>ระบบส่ง Email ทำงานปกติ &#x2705;</p>')
            flash('ส่ง Email ทดสอบสำเร็จ' if ok else 'ส่งไม่สำเร็จ กรุณาตรวจสอบการตั้งค่า',
                  'success' if ok else 'danger')

        elif action == 'send_overdue':
            overdue = db_fetchall(conn, '''
                SELECT b.*, e.name as eq_name, u.name as user_name, u.email as user_email
                FROM borrows b
                JOIN equipment e ON b.equipment_id=e.id
                JOIN users u ON b.user_id=u.id
                WHERE b.status='borrowed' AND b.due_date IS NOT NULL AND b.due_date < CURRENT_DATE
            ''')
            sent = sum(1 for b in overdue
                       if email_overdue(b['user_email'], b['user_name'], b['eq_name'], b['due_date']))
            flash(f'ส่ง Email แจ้งเตือนเกินกำหนด {sent}/{len(overdue)} รายการ', 'success')

        conn.close()
        return redirect(url_for('email_settings_page'))

    conn.close()
    return render_template('email_settings.html', cfg=cfg)


# ── SCAN QR ────────────────────────────────────────────────────────────────────

@app.route('/scan')
@login_required
def scan_qr():
    return render_template('scan_qr.html')


@app.route('/api/equipment/<int:eid>/qr')
@login_required
def api_qr(eid):
    conn = get_db()
    eq   = db_fetchone(conn, 'SELECT * FROM equipment WHERE id=%s', (eid,))
    conn.close()
    if not eq:
        return jsonify({'error': 'not found'}), 404
    qr_data = f"ID:{eid}|ชื่อ:{eq['name']}|Serial:{eq['serial_number']}|หมวด:{eq['category']}"
    return jsonify({'qr': generate_qr(qr_data), 'equipment': eq})


# ── OVERDUE MANAGEMENT ─────────────────────────────────────────────────────────

@app.route('/admin/overdue')
@admin_required
def overdue_list():
    today     = date.today()
    today_str = today.isoformat()
    conn      = get_db()
    overdue = db_fetchall(conn, '''
        SELECT b.*, e.name as eq_name, e.category, e.serial_number,
               u.name as user_name, u.department, u.email as user_email, u.phone as user_phone
        FROM borrows b
        JOIN equipment e ON b.equipment_id=e.id
        JOIN users u ON b.user_id=u.id
        WHERE b.status='borrowed' AND b.due_date IS NOT NULL AND b.due_date < CURRENT_DATE
        ORDER BY b.due_date ASC
    ''')
    conn.close()
    return render_template('overdue.html', overdue=overdue, today=today_str)


# ── CATEGORIES ─────────────────────────────────────────────────────────────────

@app.route('/admin/categories', methods=['GET', 'POST'])
@admin_required
def category_list():
    conn = get_db()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if name:
            try:
                db_execute(conn, 'INSERT INTO categories (name) VALUES (%s)', (name,))
                conn.commit()
                flash(f'เพิ่มหมวดหมู่ "{name}" สำเร็จ', 'success')
            except psycopg2.IntegrityError:
                conn.rollback()
                flash('หมวดหมู่นี้มีอยู่แล้ว', 'danger')
        conn.close()
        return redirect(url_for('category_list'))
    categories = db_fetchall(conn, 'SELECT * FROM categories ORDER BY name')
    conn.close()
    return render_template('categories.html', categories=categories)


@app.route('/admin/categories/<int:cid>/delete', methods=['POST'])
@admin_required
def delete_category(cid):
    conn = get_db()
    cat  = db_fetchone(conn, 'SELECT * FROM categories WHERE id=%s', (cid,))
    if cat:
        in_use = db_fetchone(conn, 'SELECT COUNT(*) as c FROM equipment WHERE category=%s', (cat['name'],))
        if in_use['c'] > 0:
            flash(f'ไม่สามารถลบได้ มีอุปกรณ์ใช้หมวดหมู่นี้อยู่ {in_use["c"]} รายการ', 'danger')
        else:
            db_execute(conn, 'DELETE FROM categories WHERE id=%s', (cid,))
            conn.commit()
            flash(f'ลบหมวดหมู่ "{cat["name"]}" สำเร็จ', 'success')
    conn.close()
    return redirect(url_for('category_list'))


# ── SERVE UPLOADS ──────────────────────────────────────────────────────────────

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


init_db()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
